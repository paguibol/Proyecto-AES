import subprocess
from time import sleep
from datetime import datetime
import os
import sys
import openpyxl  #Libreria para manejar archivos Excel, para guardar los resultados de las pruebas
from openpyxl import Workbook
from menu_AES import build_filename_all, build_filename


# Mode-specific defaults. AES_MODE controls which set is active in a child process.
DEFAULTS_BY_MODE = {
    "SIN_SALDO": {
        "MMS_REPS":       5,
        "MMS_INTERVAL":   60,
        "INTERNET_REPS":  5,
        "INTERNET_INTERVAL": 60,
        "YT_REPS":        5,
        "YT_INTERVAL":    60,
        "GMAIL_REPS":     5,
        "GMAIL_INTERVAL": 60,
        "TWITTER_REPS":   2,
        "TWITTER_INTERVAL": 60,
        "MESSENGER_REPS": 5,
        "MESSENGER_INTERVAL": 60,
    },
    "PREPAGO": {
        "MMS_REPS":       20,
        "MMS_INTERVAL":   60,
        "INTERNET_REPS":  10,
        "INTERNET_INTERVAL": 60,
        "YT_REPS":        10,
        "YT_INTERVAL":    60,
        "GMAIL_REPS":     20,
        "GMAIL_INTERVAL": 60,
        "TWITTER_REPS":   20,
        "TWITTER_INTERVAL": 60,
        "MESSENGER_REPS": 20,
        "MESSENGER_INTERVAL": 60,
    },
}

# POSPAGO same as PREPAGO by default
DEFAULTS_BY_MODE["POSPAGO"] = DEFAULTS_BY_MODE["PREPAGO"]

# Select active defaults based on AES_MODE env var at import time.
_mode = os.environ.get("AES_MODE", "SIN_SALDO").upper()
DEFAULTS = DEFAULTS_BY_MODE.get(_mode, DEFAULTS_BY_MODE["SIN_SALDO"]) 


def get_cfg(key):
    try:
        # environment variable overrides active defaults
        return int(os.environ.get(key, DEFAULTS[key]))
    except Exception:
        return DEFAULTS.get(key)

def adb(command):
    result = subprocess.run(["adb"] + command.split(), capture_output=True, text=True)
    return result.stdout.strip()


def connection():
    devices = adb("devices").splitlines()
    connected = [line for line in devices if "\tdevice" in line]
    device_name = adb("shell getprop ro.product.model")
    if not connected:
        print("No devices detected")
        return False
    print(f"Device detected: {device_name} ")
    return True


def go_home(d):
    d.press("home")
    sleep(2)

def go_back(d):
    d.press("back")
    sleep(2)


def open_bbklogs(d):
    d.app_start("com.google.android.dialer")
    dialpad_btns = [
        d(resourceId="com.google.android.dialer:id/dialpad_fab"),
        d(resourceId="com.google.android.dialer:id/tab_dialpad"),
        d(description="Teclado"),
        d(description="Dialpad"),
    ]
    found = None
    for sel in dialpad_btns:
        if sel.wait(timeout=5):
            found = sel
            break
    if not found:
        print("Dialer: dialpad button not found")
        return
    info = found.info or {}
    bounds = info.get("bounds")
    if bounds:
        x = (bounds["left"] + bounds["right"]) // 2
        y = (bounds["top"] + bounds["bottom"]) // 2
        d.click(x, y)
    else:
        found.click()
    sleep(4)
    if not d(className="android.widget.EditText").wait(timeout=10):
        return
    d(className="android.widget.EditText").set_text("*#*#112#*#*")
    sleep(2)
    d.app_stop("com.google.android.dialer")

def take_screenshot(d, prefix="screenshot"):
    # Determine technology and network from environment or script name
    script_base = None
    try:
        script_base = os.path.splitext(os.path.basename(sys.argv[0]))[0]
    except Exception:
        script_base = prefix

    # Known test names (prefer nicer capitalization)
    tests_map = {
        "mms": "MMS",
        "internet": "Internet",
        "youtube": "Youtube",
        "twitter": "Twitter",
        "messenger": "Messenger",
        "gmail": "Gmail",
    }

    tech = None
    low = (script_base or "").lower()
    for k, v in tests_map.items():
        if k in low:
            tech = v
            break

    # Fallback to environment-provided values
    env_mode = os.environ.get("AES_MODE", "").replace("_", " ").strip()
    env_network = os.environ.get("AES_NETWORK", "").strip()

    # detect network from script name if not in env
    network = env_network or ("4G" if "4g" in low else ("3G" if "3g" in low else ""))

    if not tech and env_mode:
        tech = env_mode.title()

    if not tech:
        tech = prefix

    date_str = datetime.now().strftime('%Y-%m-%d')
    time_str = datetime.now().strftime('%H-%M-%S')

    # If running in SIN_SALDO mode, group all screenshots for that network in a common folder
    env_mode_raw = os.environ.get("AES_MODE", "").upper()
    if env_mode_raw in ("SIN_SALDO", "SIN SALDO", "SIN-SALDO"):
        net_label = network or ""
        folder = os.path.join("logs", f"Screenshots Sin Saldo {net_label} {date_str}")
        os.makedirs(folder, exist_ok=True)
        safe_label = f"{tech}_{net_label}".strip().replace(" ", "_")
        filename = f"{safe_label}_{date_str}_{time_str}.png"
    else:
        label = f"{tech} {network}".strip()
        folder = os.path.join("logs", f"Screenshots {label} {date_str}")
        os.makedirs(folder, exist_ok=True)
        # filename: use underscore-separated label + timestamp
        safe_label = label.replace(" ", "_") if label else prefix
        filename = f"{safe_label}_{date_str}_{time_str}.png"
    path = os.path.join(folder, filename)
    try:
        d.screenshot(path)
        print(f"Screenshot saved: {path}")
    except Exception as e:
        print("Failed to take screenshot:", e)



######### Función, guardar la hora en Excel, en la hoja "Señalización_3G" o "Señalización_4G" para 2 columnas (MMS, FB, X, ...)
# Ejemplo de llamar a la funcion en cada programa: write_time_to_Excel_2_columns(i+1, current_time, col_a="C", col_b="D", start_row=36, NW="4G")    # col_a =columna a iniciar ;fila; pestaña del excel en cual tecnologia Señalización_3G o ...4G 


def write_time_to_Excel_2_columns(iteration, timestamp, col_a, col_b, start_row, NW, total_reps):
    #checar ruta, yo tengo otra version
    if getattr(sys, 'frozen', False):                             # Para guardar en excel en donde esta el ejecutable .exe
        ruta_base = os.path.dirname(sys.executable)               # Si el programa está congelado, usa la ruta del ejecutable.
    else:
        ruta_base = os.path.dirname(os.path.abspath(__file__))    # Si no, usa la ruta del script .py



    #ruta_base = os.environ.get("AES_RUTA_BASE", os.path.dirname(os.path.abspath(__file__)))
    nombre_archivo = "Pruebas_Homologacion_PS_3G_4G.xlsx"
    path_completo = os.path.join(ruta_base, nombre_archivo)

    try:
        wb = openpyxl.load_workbook(path_completo)
        
        nombre_hoja = f"Señalización_{NW}"
        
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
        else:
            print(f"Error: No se encontró la pestaña (no escribió 3G o 4G en variable NW) '{nombre_hoja}'")
            return

        fila_destino = start_row + ((iteration - 1) // 2)         # Calcular fila: aumenta cada 2 iteraciones.        Iter 1,2 -> fila inicial | Iter 3,4 -> fila inicial + 1...

        
        if iteration % 2 != 0:      #  Elegir columna:         Si la iteración es IMPAR (1, 3, 5...), usa la primera columna solicitada
            col_final = col_a
        
        else:                       # Si la iteración es PAR (2, 4, 6...), usa la segunda columna solicitada
            col_final = col_b

        
        ws[f"{col_final}{fila_destino}"] = timestamp    # Escritura
        wb.save(path_completo)
        print(f"Guardado en {col_final}{fila_destino}: {timestamp}")




        # --- LÓGICA DE RETORNO HORA INICIO Y FIN ---
        if iteration == 1:           # Si es la primera iteración, regresamos el timestamp para guardarlo fuera
            return "PRIMERA", timestamp
        
        elif iteration == total_reps:      # Si es la última iteración, regresamos el timestamp
            return "ULTIMA", timestamp
        
        return None                        # Si es una iteración intermedia, regresa None (nada) 
    


    
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{nombre_archivo}' en la carpeta: {ruta_base}")
    except PermissionError:
        print(f"Error: No se pudo guardar. Cierra el archivo Excel antes de ejecutar.")
    
    except Exception as e:
        print(f"Error al acceder al Excel, cierre el excel y guardelo con el nombre correcto en la carpeta - Pruebas_Homologacion_PS_3G_4G : {e}")

######### Función, guardar la hora en Excel, en la hoja "Señalización_3G" o "Señalización_4G"para 1 columnas (Internet, youtube, MMS (prepago sin balance), ...)
# Ejemplo de llamar a la funcion en cada programa:     write_time_to_Excel_1_column(i+1, current_time, col="T", start_row=36, NW="3G")


def write_time_to_Excel_1_column(iteration, timestamp, col, start_row, NW, total_reps):
    #Chechar ruta, yo tengo otra version
    if getattr(sys, 'frozen', False):                             # Para guardar en excel en donde esta el ejecutable .exe
        ruta_base = os.path.dirname(sys.executable)               # Si el programa está congelado, usa la ruta del ejecutable.
    else:
        ruta_base = os.path.dirname(os.path.abspath(__file__))    # Si no, usa la ruta del script .py


    
    #ruta_base = os.environ.get("AES_RUTA_BASE", os.path.dirname(os.path.abspath(__file__)))
    nombre_archivo = "Pruebas_Homologacion_PS_3G_4G.xlsx"
    path_completo = os.path.join(ruta_base, nombre_archivo)

    try:
        wb = openpyxl.load_workbook(path_completo)
        nombre_hoja = f"Señalización_{NW}"
        
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
        else:
            print(f"Error: No se encontró la pestaña '{nombre_hoja}' (verifique variable NW)")
            return

        # --- LÓGICA DE COLUMNA ÚNICA ---       
        fila_destino = start_row + (iteration - 1)          # La fila aumenta simplemente sumando la iteración actual        # Iter 1 -> start_row + 0
        col_final = col                                     # La columna es fija, 
       
        ws[f"{col_final}{fila_destino}"] = timestamp        # Escritura
        wb.save(path_completo)
        print(f"[{nombre_hoja}] Guardado en {col_final}{fila_destino}: {timestamp}")



        # --- NUEVA LÓGICA DE RETORNO ---
        if iteration == 1:           # Si es la primera iteración, regresamos el timestamp para guardarlo fuera
            return "PRIMERA", timestamp
        
        elif iteration == total_reps:      # Si es la última iteración, regresamos el timestamp
            return "ULTIMA", timestamp
        
        return None                        # Si es una iteración intermedia, regresa None (nada) 





    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{nombre_archivo}' en la carpeta: {ruta_base}")
    except PermissionError:
        print(f"Error: No se pudo guardar. Cierra el archivo Excel antes de ejecutar.")

    except Exception as e:
        print(f"Error al acceder al Excel (Cierre el excel o guardelo en la carpeta con el nombre correcto - Pruebas_Homologacion_PS_3G_4G ): {e}")





def write_start_end_time_test_to_Excel(tiempo_inicio, tiempo_fin, col_c, col_d, start_row, NW):  
    
    
    if getattr(sys, 'frozen', False):                             # Para guardar en excel en donde esta el ejecutable .exe
        ruta_base = os.path.dirname(sys.executable)               # Si el programa está congelado, usa la ruta del ejecutable.
    else:
        ruta_base = os.path.dirname(os.path.abspath(__file__))    # Si no, usa la ruta del script .py

    nombre_archivo = "Pruebas_Homologacion_PS_3G_4G.xlsx"
    path_completo = os.path.join(ruta_base, nombre_archivo)

    try:
        wb = openpyxl.load_workbook(path_completo)
        nombre_hoja = f"Señalización_{NW}"
        
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
        else:
            print(f"Error: No se encontró la pestaña '{nombre_hoja}' (verifique variable NW)")
            return

        # Escribir tiempo de inicio y fin en las columnas C y D respectivamente, en la fila 36 (o la que corresponda)
        ws[f"{col_c}{start_row}"] = tiempo_inicio
        ws[f"{col_d}{start_row}"] = tiempo_fin
        wb.save(path_completo)
        #print(f"[{nombre_hoja}] Guardado tiempo de inicio en {col_c}{start_row}: {tiempo_inicio}")
        #print(f"[{nombre_hoja}] Guardado tiempo de fin en {col_d}{start_row}: {tiempo_fin}")

    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{nombre_archivo}' en la carpeta: {ruta_base}")
    except PermissionError:
        print(f"Error: No se pudo guardar. Cierra el archivo Excel antes de ejecutar.")

    except Exception as e:
        print(f"Error al acceder al Excel (Cierre el excel o guardelo en la carpeta con el nombre correcto - Pruebas_Homologacion_PS_3G_4G ): {e}")



#Función para llenar el Excel con información del modelo, fecha y posteriormente el numero de la SIM

def fill_excel_with_basic_info(NW, SIM_number, Linea):  
    

    device_name = adb("shell getprop ro.product.model") #obtener el nombre del dispositivo para escribirlo en el Excel

    current_date = datetime.now().strftime('%d/%m/%Y') # obtener el tiempo actual para escribirlo en el Excel


    
    # Llenar información básica de la prueba en el Excel, como la tecnología y el número de repeticiones
    if getattr(sys, 'frozen', False):                             # Para guardar en excel en donde esta el ejecutable .exe
        ruta_base = os.path.dirname(sys.executable)               # Si el programa está congelado, usa la ruta del ejecutable.
    else:
        ruta_base = os.path.dirname(os.path.abspath(__file__))    # Si no, usa la ruta del script .py

    nombre_archivo = "Pruebas_Homologacion_PS_3G_4G.xlsx"
    path_completo = os.path.join(ruta_base, nombre_archivo)


    try:
        wb = openpyxl.load_workbook(path_completo)
        nombre_hoja = f"Señalización_{NW}"
        
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
        else:
            print(f"Error: No se encontró la pestaña '{nombre_hoja}' (verifique variable NW)")
            return

        ws["B15"] = f"Modelo: {device_name}"#Escribir el Modelo en la celda B15
        
        ws["B16"] = f"Fecha de pruebas: {current_date}" # Escribir la Fecha de pruebas en la celda B16

        if Linea == "Pospago":
            ws["F14"] = f"Línea Pospago: {SIM_number}"
        elif Linea == "Prepago sin saldo":
            ws["F15"] = f"Línea Prepago con saldo: {SIM_number}"
        elif Linea == "Prepago con saldo":
            ws["F16"] = f"Línea Prepago sin saldo: {SIM_number}"
        else:
            print("Error: Linea debe ser 'Pospago' o 'Prepago sin saldo' o 'Prepago con saldo' ")
        

        wb.save(path_completo)





    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{nombre_archivo}' en la carpeta: {ruta_base}")
    except PermissionError:
        print(f"Error: No se pudo guardar. Cierra el archivo Excel antes de ejecutar.")

    except Exception as e:
        print(f"Error al acceder al Excel (Cierre el excel o guardelo en la carpeta con el nombre correcto - Pruebas_Homologacion_PS_3G_4G ): {e}")




def get_number_SIM(d):
    d.app_start("com.android.settings")
    sleep(2)
    d(resourceId="android:id/title", text="Red móvil").click_exists(timeout=5)
    sleep(2)
    d(resourceId="com.android.phone:id/title", text="SIM 1").click_exists(timeout=5)
    sleep(2)
    d(resourceId="com.android.phone:id/title", text="Número de la tarjeta SIM").click_exists(timeout=5)
    sleep(2)
    complete_text = d(resourceId="com.android.phone:id/edit").get_text()
    number_10_digits = complete_text[3:]
    d.app_stop("com.android.settings")
    #print(f"SIM number: {number_10_digits}")
    d.app_stop("com.android.settings")
    return number_10_digits
